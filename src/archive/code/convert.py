import pytesseract
import openpyxl
from openpyxl import load_workbook

workbook = load_workbook('output1.xlsx')
worksheet = workbook.get_active_sheet()

html_data = """
 <html>
    <head>
        <title>
       Table
        </title>
        <style>
        table, tr, td {
        border: 0px solid black;
        }
</style>
    </head>
    <body>
       
    <table style="border: solid 1px #000000">
    """

ws_range = worksheet.iter_rows('A1:H13')
for row in ws_range:
    html_data += "<tr>"
    for cell in row:
        if cell.value is None:
            html_data += "<td>" + ' ' + "</td>"
        else:
            html_data += "<td>" + str(cell.value) + "</td>"
    html_data += "</tr>"
html_data += "</table></body></html>"

with open("output.html", "w") as html_fil:
    html_fil.write(html_data)

