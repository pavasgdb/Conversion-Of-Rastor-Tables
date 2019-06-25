import os
from tabula import wrapper
import openpyxl
from openpyxl import load_workbook
tables = wrapper.read_pdf("ncert.pdf",multiple_tables=True,pages='94,95,96')

i=1
for table in tables:
    table.to_excel('output'+str(i)+'.xlsx',index=False)
    print(i)
    i=i+1
    workbook = load_workbook('output'+str(i)+'.xlsx')
    worksheet = workbook.get_active_sheet()

    html_data = """
    <html>
    <head>
        <title>
       Table
        </title>
        <style>
        table, tr, td {
        border: 0px solid black;
        }
</style>
    </head>
    <body>
       
    <table style="border: solid 1px #000000">
    """

    ws_range = worksheet.iter_rows('A1:H13')
    for row in ws_range:
        html_data += "<tr>"
        for cell in row:
            if cell.value is None:
                html_data += "<td>" + ' ' + "</td>"
            else:
                html_data += "<td>" + str(cell.value) + "</td>"
        html_data += "</tr>"
    html_data += "</table></body></html>"

    with open("output_1_"+str(i)+".html", "w") as html_fil:
        html_fil.write(html_data)

    
